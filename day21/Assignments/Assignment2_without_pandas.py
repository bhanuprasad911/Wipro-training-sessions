from openpyxl import load_workbook, Workbook
wb = load_workbook("sales_data.xlsx")
ws = wb["2025"]
ws["D1"] = "Total"

for row in range(2, ws.max_row + 1):
    quantity = ws[f"B{row}"].value
    price = ws[f"C{row}"].value
    ws[f"D{row}"] = quantity * price
wb.save("sales_summary.xlsx")
print("sales_summary.xlsx created successfully!")